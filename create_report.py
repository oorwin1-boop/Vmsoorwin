import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── colour palette ──────────────────────────────────────────────
HDR_BG   = "1F3864"   # dark navy
HDR_FG   = "FFFFFF"
PASS_BG  = "C6EFCE"   # light green
PASS_FG  = "276221"
FAIL_BG  = "FFE0E0"   # light red
FAIL_FG  = "9C0006"
WARN_BG  = "FFEB9C"   # light orange
WARN_FG  = "9C6500"
INFO_BG  = "DDEBF7"   # light blue
INFO_FG  = "1F4E79"
ALT_BG   = "F2F2F2"   # alternating row grey

thin = Side(style="thin", color="BBBBBB")
med  = Side(style="medium", color="888888")
def brd(top=thin, bot=thin, left=thin, right=thin):
    return Border(top=top, bottom=bot, left=left, right=right)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=HDR_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = brd(top=med, bot=med, left=med, right=med)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def status_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    v = value.upper() if value else ""
    if v == "PASS":
        c.fill, c.font = PatternFill("solid", fgColor=PASS_BG), Font(bold=True, color=PASS_FG)
    elif v in ("FAIL", "BUG"):
        c.fill, c.font = PatternFill("solid", fgColor=FAIL_BG), Font(bold=True, color=FAIL_FG)
    elif v in ("WARN", "WARNING", "OBSERVATION"):
        c.fill, c.font = PatternFill("solid", fgColor=WARN_BG), Font(bold=True, color=WARN_FG)
    elif v in ("INFO", "N/A"):
        c.fill, c.font = PatternFill("solid", fgColor=INFO_BG), Font(bold=True, color=INFO_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = brd()

def data_cell(ws, row, col, value, alt=False, bold=False, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    if alt:
        c.fill = PatternFill("solid", fgColor=ALT_BG)
    c.font      = Font(bold=bold, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border    = brd()
    return c

# ════════════════════════════════════════════════════════════════
# SHEET 1 — Test Execution Summary
# ════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Test Execution Summary"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

# Title banner
ws1.merge_cells("A1:H1")
t = ws1["A1"]
t.value = "BeanHire ATS — Job Posting Module  |  Test Execution Report"
t.font  = Font(bold=True, size=14, color=HDR_FG)
t.fill  = PatternFill("solid", fgColor=HDR_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:H2")
sub = ws1["A2"]
sub.value = f"Environment: oorwinlabs.beanhiredev.com  |  User: pavant+ui@oorwin.com  |  Date: 2026-06-11  |  Tester: Claude AI"
sub.font  = Font(italic=True, size=10, color="444444")
sub.fill  = PatternFill("solid", fgColor="D9E1F2")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

# Column headers
headers = ["#", "Module", "Test Case", "Steps Summary", "Expected Result",
           "Actual Result", "Status", "OOPS Error?"]
widths  = [4,  18,  32,  42,  30,  40,  12,  14]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    hdr_cell(ws1, 3, i, h, w)
ws1.row_dimensions[3].height = 28

# Test data
rows = [
    # (module, test_case, steps, expected, actual, status, oops)
    ("Login",
     "Valid login with correct credentials",
     "1. Navigate to https://oorwinlabs.beanhiredev.com\n2. Enter email: pavant+ui@oorwin.com\n3. Enter password\n4. Click Sign In",
     "Dashboard loads successfully",
     "Dashboard loaded. Redirected to /#/hire/dashboard. User 'Pavan Kumar DEV' shown.",
     "PASS", "NO"),

    ("Jobs List",
     "Jobs list page loads with all records",
     "1. Click Recruit > Jobs in left nav\n2. Observe list, columns, pagination",
     "Jobs list shows records with all columns and pagination",
     "Jobs list loaded. 7,980 total records. 20 per page shown. Pagination controls visible. All columns (Job ID, Title, Status, Customer, etc.) present.",
     "PASS", "NO"),

    ("Jobs List",
     "Add Job dropdown options visible",
     "1. Click 'Add Job' button\n2. Observe dropdown menu",
     "Dropdown shows: Manual, Parse Text, Generate JD & Create Job, Upload & Parse Document",
     "All 4 options displayed correctly.",
     "PASS", "NO"),

    ("Create Job",
     "Create Job form loads (Manual)",
     "1. Click Add Job > Manual\n2. Observe form sections",
     "Create Job form opens with all sections: Job Details, Pay & Billing, Recruitment Team, Job Description",
     "Form loaded at /#/jobs/create with all required sections visible.",
     "PASS", "NO"),

    ("Create Job",
     "Save without filling required fields — validation",
     "1. Open Create Job form\n2. Click Save without entering any data",
     "Error messages appear on each required field guiding the user",
     "Form has 30+ ng-invalid fields marked with red borders but NO visible error messages are rendered. User cannot identify which fields are required.",
     "BUG", "NO"),

    ("Create Job",
     "Duplicate 'Client Job ID' field in form",
     "1. Open Create Job form\n2. Scroll through Job Details section",
     "Each field should appear only once",
     "'Client Job ID' field appears TWICE in the Job Details section (two separate input fields with the same label).",
     "BUG", "NO"),

    ("Create Job",
     "Spelling error: 'Mark as Preffered'",
     "1. Open Create Job form\n2. View Pay & Billing Details section",
     "Checkbox label reads 'Mark as Preferred'",
     "Checkbox label reads 'Mark as Preffered' (double 'r') — spelling mistake.",
     "BUG", "NO"),

    ("Create Job",
     "Test/dev fields visible in production form",
     "1. Open Create Job form\n2. Scroll through all fields",
     "Only production-ready fields should be visible",
     "Multiple test/dev fields visible: 'new test sr', 'max input', 'Length Field', 'Text 256', 'Test 256 1', 'Test 256 5', 'Test 256 6', 'test bill', 'tech Skill', 'Q3 Qualified Jobs', 'q4 jobs test', 'q5 jobs test', 'Q4 Qualified Jobs'. Section named 'Text Field Validation' also visible.",
     "BUG", "NO"),

    ("Job Detail",
     "View existing job detail — all sections load",
     "1. From Jobs list, click on job OOJ-86071\n2. Observe all tabs and sections",
     "Job detail opens with all sections expanded and correct data",
     "Job detail loaded at /#/jobs/9554/info. Sections visible: Job Details, Text Field Validation, Pay & Billing Details, Recruitment Team, Recruiter Instructions, Job Description, Attachments, Tags. Stats bar shows Pipeline/Submitted/End client/Interviews/Confirmations/Rejected/Onboarded counts.",
     "PASS", "NO"),

    ("Job Detail",
     "All detail tabs navigate correctly",
     "1. Click each tab: Overview, Info, Pipeline, Submissions, Interviews, Communications, Recruiter Actions, Assessments",
     "Each tab loads without errors",
     "All 8 tabs loaded correctly. No navigation errors. URLs updated correctly for each tab.",
     "PASS", "NO"),

    ("Edit Job",
     "Edit existing job and save",
     "1. Open Job Detail\n2. Click Edit button\n3. Form loads with existing data\n4. Click Update",
     "Edit form pre-populated with existing values. After Update, redirects to Job Info view.",
     "Edit form loaded at /#/jobs/edit/9554 with all fields pre-filled (Job Title: 'Java developer', Bill Rate: 3–4, Pay Rate: 2–6, Customer: 'aws16'). Clicking Update redirected back to /#/jobs/9554/info successfully.",
     "PASS", "NO"),

    ("Clone Job",
     "Clone existing job",
     "1. Open Job Detail\n2. Click Clone button",
     "Clone form opens pre-filled with original job data",
     "Clone form loaded at /#/jobs/clone/9554/1 with page title 'Clone Job'. Form pre-populated with source job data.",
     "PASS", "NO"),

    ("Publish",
     "Publish job to Job Boards & Career Portal",
     "1. Open Job Detail\n2. Click Publish > Job Boards & Career Portal",
     "Publish dialog opens showing available job boards",
     "Publish dialog opened. Job boards listed: Career Portal, Jobserve (Premium), jobboard_com (Premium), Indeed, Post Job Free, Glassdoor and more. No OOPS error.",
     "PASS", "NO"),

    ("Share",
     "Share job via Share button",
     "1. Open Job Detail\n2. Click Share (share icon) button",
     "Share panel opens with sharing options",
     "Share panel appeared with 3 options: Share Job, Share Job Via SMS, Compose Mail. Panel dismissed correctly via Escape.",
     "PASS", "NO"),

    ("Console Errors",
     "Browser console errors during session",
     "Monitor browser console throughout all test steps",
     "0 console errors expected",
     "25 console errors accumulated across the session (started at 1 on login, grew to 25 by Communications tab). These are background JS errors. No visible impact on UI in positive flows but may indicate API call failures or component issues.",
     "WARN", "NO"),
]

for idx, (module, tc, steps, expected, actual, status, oops) in enumerate(rows, 1):
    r = idx + 3
    alt = (idx % 2 == 0)
    ws1.row_dimensions[r].height = 60
    data_cell(ws1, r, 1, idx, alt, bold=True, wrap=False)
    data_cell(ws1, r, 2, module, alt, bold=True)
    data_cell(ws1, r, 3, tc, alt)
    data_cell(ws1, r, 4, steps, alt)
    data_cell(ws1, r, 5, expected, alt)
    data_cell(ws1, r, 6, actual, alt)
    status_cell(ws1, r, 7, status)
    c = ws1.cell(row=r, column=8, value=oops)
    c.fill = PatternFill("solid", fgColor=FAIL_BG if oops == "YES" else PASS_BG)
    c.font = Font(bold=True, color=FAIL_FG if oops == "YES" else PASS_FG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = brd()

# ════════════════════════════════════════════════════════════════
# SHEET 2 — Bug Report
# ════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Bug Report")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A4"

ws2.merge_cells("A1:I1")
t2 = ws2["A1"]
t2.value = "BeanHire ATS — Job Posting Module  |  Bug Report"
t2.font  = Font(bold=True, size=14, color=HDR_FG)
t2.fill  = PatternFill("solid", fgColor="7B2C2C")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:I2")
sub2 = ws2["A2"]
sub2.value = "Bugs and observations found during positive functional testing of the Job Posting module — 2026-06-11"
sub2.font  = Font(italic=True, size=10)
sub2.fill  = PatternFill("solid", fgColor="F4CCCC")
sub2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 20

bug_headers = ["#", "Module", "Severity", "Bug Title", "Steps to Reproduce",
               "Expected", "Actual / Impact", "OOPS?", "Screenshot"]
bug_widths  = [4,   16,   12,  34,   42,   28,   40,   10,   22]
for i, (h, w) in enumerate(zip(bug_headers, bug_widths), 1):
    hdr_cell(ws2, 3, i, h, w)
ws2.row_dimensions[3].height = 28

bugs = [
    ("Create Job", "HIGH",
     "No error messages on required-field validation",
     "1. Navigate to Create Job (Manual)\n2. Click Save without filling any fields",
     "Each required field shows an inline error message (e.g. 'This field is required')",
     "Form fields turn red (ng-invalid) but NO error messages are rendered. User cannot know which fields are required or what to fix. 30+ fields are silently invalid.",
     "NO", "screenshots/05b_invalid_fields_no_errors.png"),

    ("Create Job", "MEDIUM",
     "Duplicate 'Client Job ID' field in Job Details",
     "1. Open Create Job > Manual\n2. Scroll through Job Details section",
     "Field 'Client Job ID' appears exactly once",
     "Two separate 'Client Job ID' input fields exist in the same section, causing data entry confusion and potential data integrity issues.",
     "NO", "screenshots/04_create_job_blank.png"),

    ("Create Job", "LOW",
     "Spelling error: 'Mark as Preffered' checkbox",
     "1. Open Create Job > Manual\n2. View Pay & Billing Details section",
     "Label reads 'Mark as Preferred'",
     "Label reads 'Mark as Preffered' (extra 'r'). UI quality / branding issue.",
     "NO", "screenshots/04_create_job_blank.png"),

    ("Create Job", "HIGH",
     "Test/dev fields and section name visible in production",
     "1. Open Create Job > Manual\n2. Scroll through all form fields",
     "Only production-ready, user-facing fields visible",
     "14+ test fields visible: 'new test sr', 'max input', 'Length Field', 'Text 256', 'Test 256 1–6', 'test bill', 'tech Skill', 'Q3/Q4 Qualified Jobs', 'q4/q5 jobs test'. Section header 'Text Field Validation' also exposed. These appear in both Create and View/Edit forms.",
     "NO", "screenshots/04_create_job_blank.png"),

    ("Console", "MEDIUM",
     "25 JavaScript console errors accumulated during session",
     "1. Open browser DevTools console\n2. Navigate through the Job Posting module",
     "0 console errors during normal usage",
     "25 errors logged by session end. Errors began on page load and grew with each navigation. May indicate failed API calls, missing assets, or unhandled Angular exceptions. No visible UI impact in positive flows.",
     "NO", "N/A"),
]

for idx, (module, sev, title, steps, expected, actual, oops, ss) in enumerate(bugs, 1):
    r = idx + 3
    alt = (idx % 2 == 0)
    ws2.row_dimensions[r].height = 80
    data_cell(ws2, r, 1, idx, alt, bold=True, wrap=False)
    data_cell(ws2, r, 2, module, alt, bold=True)
    # Severity colour
    sc = ws2.cell(row=r, column=3, value=sev)
    sev_colors = {"HIGH": ("FFD7D7","9C0006"), "MEDIUM": ("FFEB9C","9C6500"), "LOW": ("E2EFDA","375623")}
    bg, fg = sev_colors.get(sev, (ALT_BG, "000000"))
    sc.fill = PatternFill("solid", fgColor=bg)
    sc.font = Font(bold=True, color=fg)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    sc.border = brd()
    data_cell(ws2, r, 4, title, alt, bold=True)
    data_cell(ws2, r, 5, steps, alt)
    data_cell(ws2, r, 6, expected, alt)
    data_cell(ws2, r, 7, actual, alt)
    oc = ws2.cell(row=r, column=8, value=oops)
    oc.fill = PatternFill("solid", fgColor=FAIL_BG if oops=="YES" else PASS_BG)
    oc.font = Font(bold=True, color=FAIL_FG if oops=="YES" else PASS_FG)
    oc.alignment = Alignment(horizontal="center", vertical="center")
    oc.border = brd()
    data_cell(ws2, r, 9, ss, alt)

# ════════════════════════════════════════════════════════════════
# SHEET 3 — Summary Dashboard
# ════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Summary")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:D1")
ts = ws3["A1"]
ts.value = "Test Summary"
ts.font  = Font(bold=True, size=16, color=HDR_FG)
ts.fill  = PatternFill("solid", fgColor=HDR_BG)
ts.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 40

metrics = [
    ("Test Date",          "2026-06-11"),
    ("Application",        "BeanHire ATS — oorwinlabs.beanhiredev.com"),
    ("Module Tested",      "Job Posting Module"),
    ("Test Type",          "Functional — Positive Cases"),
    ("Tester",             "Claude AI (pavant+ui@oorwin.com)"),
    ("Total Test Cases",   15),
    ("PASS",               10),
    ("FAIL / BUG",         4),
    ("WARN / Observation", 1),
    ("OOPS Errors Found",  0),
    ("Console Errors",     25),
    ("Screenshots Taken",  19),
]

for i, (label, value) in enumerate(metrics, 3):
    ws3.row_dimensions[i].height = 22
    lc = ws3.cell(row=i, column=1, value=label)
    lc.font = Font(bold=True, size=11)
    lc.fill = PatternFill("solid", fgColor="D9E1F2")
    lc.alignment = Alignment(vertical="center")
    lc.border = brd()
    ws3.column_dimensions["A"].width = 28

    vc = ws3.cell(row=i, column=2, value=value)
    vc.font = Font(size=11)
    vc.alignment = Alignment(vertical="center")
    vc.border = brd()
    ws3.column_dimensions["B"].width = 50

    # colour key metrics
    if label == "PASS":
        vc.fill = PatternFill("solid", fgColor=PASS_BG); vc.font = Font(bold=True, color=PASS_FG, size=11)
    elif label in ("FAIL / BUG",):
        vc.fill = PatternFill("solid", fgColor=FAIL_BG); vc.font = Font(bold=True, color=FAIL_FG, size=11)
    elif label in ("WARN / Observation",):
        vc.fill = PatternFill("solid", fgColor=WARN_BG); vc.font = Font(bold=True, color=WARN_FG, size=11)
    elif label == "OOPS Errors Found":
        vc.fill = PatternFill("solid", fgColor=PASS_BG); vc.font = Font(bold=True, color=PASS_FG, size=11)
    elif label == "Console Errors":
        vc.fill = PatternFill("solid", fgColor=WARN_BG); vc.font = Font(bold=True, color=WARN_FG, size=11)

ws3.merge_cells("A15:D15")
note = ws3["A15"]
note.value = (
    "KEY FINDINGS:\n"
    "✓ All positive functional flows passed (Login, View, Edit, Clone, Publish, Share, all tabs)\n"
    "✗ No OOPS errors encountered during entire test session\n"
    "⚠ 4 bugs identified: missing validation messages, duplicate field, typo, test fields in production\n"
    "⚠ 25 JavaScript console errors — requires developer investigation"
)
note.font = Font(size=10)
note.alignment = Alignment(wrap_text=True, vertical="top")
note.fill = PatternFill("solid", fgColor="FFF2CC")
note.border = Border(top=med, bottom=med, left=med, right=med)
ws3.row_dimensions[15].height = 100
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 20

out_path = r"C:\Oorwin_AI\OorwinLogin.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
