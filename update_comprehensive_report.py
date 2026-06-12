import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Oorwin_AI\OorwinLogin.xlsx')

HDR_BG  = "1F3864"; HDR_FG  = "FFFFFF"
PASS_BG = "C6EFCE"; PASS_FG = "276221"
FAIL_BG = "FFE0E0"; FAIL_FG = "9C0006"
WARN_BG = "FFEB9C"; WARN_FG = "9C6500"
INFO_BG = "DDEBF7"; INFO_FG = "1F4E79"
ALT_BG  = "F2F2F2"

thin = Side(style="thin", color="BBBBBB")
med  = Side(style="medium", color="888888")
def brd(): return Border(top=thin, bottom=thin, left=thin, right=thin)
def mbrd(): return Border(top=med, bottom=med, left=med, right=med)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=HDR_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = mbrd()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def data_cell(ws, row, col, value, alt=False, bold=False, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    if alt: c.fill = PatternFill("solid", fgColor=ALT_BG)
    c.font      = Font(bold=bold, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    c.border    = brd()

def status_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    v = (value or "").upper()
    if v == "PASS":   c.fill, c.font = PatternFill("solid", fgColor=PASS_BG), Font(bold=True, color=PASS_FG, size=10)
    elif v == "WARN": c.fill, c.font = PatternFill("solid", fgColor=WARN_BG), Font(bold=True, color=WARN_FG, size=10)
    elif v == "FAIL": c.fill, c.font = PatternFill("solid", fgColor=FAIL_BG), Font(bold=True, color=FAIL_FG, size=10)
    elif v == "INFO": c.fill, c.font = PatternFill("solid", fgColor=INFO_BG), Font(bold=True, color=INFO_FG, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = brd()

# ── Sheet 5: Comprehensive Functionality Test ──────────────────────
if "Functionality Test" in wb.sheetnames:
    del wb["Functionality Test"]

ws = wb.create_sheet("Functionality Test")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

# Banner
ws.merge_cells("A1:J1")
t = ws["A1"]
t.value = "Job Posting Module — Comprehensive Functionality Test Results  |  Job: OOJ-86071 (ID: 9554)  |  Date: 2026-06-11"
t.font  = Font(bold=True, size=13, color=HDR_FG)
t.fill  = PatternFill("solid", fgColor=HDR_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:J2")
sub = ws["A2"]
sub.value = "Tester: pavant+ui@oorwin.com  |  Base URL: https://oorwinlabs.beanhiredev.com  |  Scope: All functional positive cases + all buttons/menus/tabs"
sub.font  = Font(italic=True, size=10)
sub.fill  = PatternFill("solid", fgColor="D9E1F2")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

headers = ["#", "Feature Area", "Action Tested", "Expected Result", "Actual Result", "OOPS?", "Dialog/Toast", "Issues Found", "Screenshot", "Status"]
widths  = [4,    22,             30,              28,               35,              8,       20,              35,             18,           10]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    hdr_cell(ws, 3, i, h, w)
ws.row_dimensions[3].height = 26

rows = [
    # (area, action, expected, actual, oops, dialog_toast, issues, screenshot, status)
    ("Add Job — Parse Text",
     "Click 'Parse Text' option from Add Job dropdown",
     "Dialog opens for text input",
     "Dialog 'Parse Text' opened with text area for job description parsing",
     "NO", "Parse Text dialog", "None", "parse_text_dialog.png", "PASS"),

    ("Add Job — Generate JD",
     "Click 'Generate JD' option from Add Job dropdown",
     "AI generates job description",
     "Dialog opened; AI generated Java developer JD with 5 sections; saved to form successfully",
     "NO", "Generate JD dialog", "None", "generate_jd_result.png", "PASS"),

    ("Add Job — Upload & Parse",
     "Click 'Upload & Parse Document' from Add Job dropdown",
     "File picker opens for document upload",
     "File chooser triggered (OS-level); no OOPS; cancelled safely",
     "NO", "OS file picker", "None", "—", "PASS"),

    ("Job Detail — Candidate dropdown",
     "Click Candidate button → view menu options",
     "Menu shows submission options",
     "Menu shows: 'Submit Candidate', 'Parse & Submit Candidate'",
     "NO", "Dropdown menu", "None", "—", "PASS"),

    ("Job Detail — Pipeline dropdown",
     "Click Pipeline button → view menu options",
     "Menu shows pipeline options",
     "Menu shows: 'Add Pipeline', 'Parse & Add Pipeline'",
     "NO", "Dropdown menu", "None", "—", "PASS"),

    ("Job Detail — Stats Bar (7 buttons)",
     "Click all 7 stat buttons: Pipeline, Submitted, End client, Interviews, Confirmations, Rejected, Onboarded",
     "Each button highlights active state",
     "All 7 buttons toggle [active] state. No navigation (counts all = 0). No errors.",
     "NO", "Visual toggle only", "None", "stats_buttons_tested.png", "PASS"),

    ("Job Detail — more_vert: Copy Job URL",
     "Open more_vert menu → click 'Copy Job URL'",
     "Job URL copied to clipboard silently",
     "Clicked; no dialog; no toast; URL silently copied to clipboard",
     "NO", "Silent clipboard copy", "None", "more_vert_menu.png", "PASS"),

    ("Job Detail — more_vert: Tag Assessment",
     "Open more_vert menu → click 'Tag Assessment'",
     "Assessment tool picker dialog opens",
     "Dialog 'Select Assessment Tool' opened with 4 providers: Adaface, Oorwin, Testlify, Codility",
     "NO", "Select Assessment Tool dialog", "None", "tag_assessment_dialog.png", "PASS"),

    ("Job Detail — more_vert: Generate AI Assessment",
     "Open more_vert menu → click 'Generate AI Assessment' → enter 5 questions → Generate",
     "AI generates assessment questions",
     "Dialog opened showing 10,000 token balance; entered 5 questions; AI generated 5 Java MCQs successfully",
     "NO", "Generate AI Assessment dialog", "None", "ai_assessment_generated.png", "PASS"),

    ("Job Detail — more_vert: Edit Campaign Fields",
     "Open more_vert menu → click 'Edit Campaign Fields'",
     "Campaign fields edit dialog opens",
     "Dialog 'Edit Campaign Fields - Java developer' opened with Source Campaigns, Follow Up Campaigns, Customer, Job Description fields",
     "NO", "Edit Campaign Fields dialog", "None", "edit_campaign_fields.png", "PASS"),

    ("Job Detail — more_vert: Timeline",
     "Open more_vert menu → click 'Timeline'",
     "Job timeline dialog opens",
     "Dialog 'Job - Timeline' opened; shows 'No Data' (no activity yet)",
     "NO", "Job - Timeline dialog", "None", "timeline_dialog.png", "PASS"),

    ("Job Detail — more_vert: Tag Spark Hire Job",
     "Open more_vert menu → click 'Tag Spark Hire Job'",
     "Spark Hire tagging dialog opens",
     "Dialog opened with 'Create New Job' and 'Tag Existing Job' radio options + Cancel/Proceed",
     "NO", "Tag Spark Hire Job dialog", "None", "tag_spark_hire_job.png", "PASS"),

    ("Communications — Synced Emails",
     "Navigate to Communications > Synced Emails",
     "Tab loads, shows email connect prompt",
     "Loaded; email account connect prompt shown; 'Proceed' link to email preferences visible",
     "NO", "N/A", "None", "comms_synced_emails.png", "PASS"),

    ("Communications — Sent Emails",
     "Navigate to Communications > Sent Emails",
     "Tab loads with sent emails table",
     "Loaded with table structure; 0 sent emails",
     "NO", "N/A", "None", "—", "PASS"),

    ("Communications — Calls",
     "Navigate to Communications > Calls",
     "Tab loads with calls table",
     "Loaded with table structure; 0 calls",
     "NO", "N/A", "None", "—", "PASS"),

    ("Communications — SMS",
     "Navigate to Communications > SMS",
     "Tab loads with SMS table",
     "Loaded with table structure; 0 SMS",
     "NO", "N/A", "None", "—", "PASS"),

    ("Communications — Job Shares",
     "Navigate to Communications > Job Shares",
     "Tab loads with job shares table",
     "Loaded with table structure; 0 job shares",
     "NO", "N/A", "None", "—", "PASS"),

    ("Communications — Follow Up Campaigns",
     "Navigate to Communications > Follow Up Campaigns",
     "Tab loads with campaigns table",
     "Loaded with table structure; 0 campaigns",
     "NO", "N/A", "None", "—", "PASS"),

    ("Communications — Bot Conversation",
     "Navigate to Communications > Bot Conversation",
     "Tab loads with bot conversation log",
     "Loaded with table structure; 0 bot conversations",
     "NO", "N/A", "None", "—", "PASS"),

    ("Communications — Spark Hire Video",
     "Navigate to Communications > Spark Hire Video Interviews",
     "Tab loads with video interview table",
     "Loaded; table with INITIATED/COMPLETED filters; 'Submit All' button; columns: NAME, EMAIL, INITIATED DATE, INITIATED BY, PIPELINE STATUS",
     "NO", "N/A", "None", "comms_sparkhire.png", "PASS"),

    ("Recruiter Actions — Track Status",
     "Click Track Status checkbox on Recruiter Actions page",
     "Checkbox toggles on/off",
     "Track Status mat-checkbox found and toggled ON (checked=true) successfully",
     "NO", "N/A", "None", "recruiter_actions.png", "PASS"),

    ("Recruiter Actions — Save",
     "Click Save button after toggling Track Status",
     "Success toast shown",
     "Toast: 'Recruiter Actions updated'; data saved; no OOPS",
     "NO", "Toast: 'Recruiter Actions updated'", "None", "recruiter_actions_saved.png", "PASS"),

    ("Recruiter Actions — Clear",
     "Click Clear button on Recruiter Actions page",
     "Job board selections cleared",
     "Clear button clicked; job board checkboxes reset; no OOPS",
     "NO", "N/A", "None", "—", "PASS"),

    ("Assessments — Tagged (Oorwin provider)",
     "Click Oorwin tab → Tag Assessments → select Java Test Basic → Tag Assessment",
     "Assessment tagged to job",
     "Oorwin tab active (tool_type=2); Tag Assessments dialog showed 20+ Oorwin assessments; Java Test Basic (4 questions) selected; tagged successfully",
     "NO", "Toast: 'Selected Job(s) tagged to selected assessment(s)'", "BUG: 'Candidatee' typo in 'Write on Behalf of Candidatee' button", "assessment_tagged_success.png", "PASS"),

    ("Assessments — Initiate Assessment",
     "Click 'Initiate Assessment' button",
     "Dialog to initiate assessment for candidate",
     "Button clicked; no dialog opened (requires a candidate submission — 0 submissions exist for this job)",
     "NO", "N/A (no candidates)", "None", "assessments_tagged.png", "INFO"),

    ("Assessments — Initiated sub-tab",
     "Navigate to Initiated Assessments sub-tab",
     "Table loads with initiated assessment records",
     "Loaded; columns: S.NO, ASSESSMENT, INITIATED ON, CANDIDATE ID, CANDIDATE NAME, TEST DURATION, NO. OF QUESTIONS; 0 records",
     "NO", "N/A", "None", "—", "PASS"),

    ("Assessments — Pending sub-tab",
     "Navigate to Pending Assessments sub-tab",
     "Table loads with pending records",
     "Loaded; columns: S.NO, ASSESSMENT, TRIGGERED ON, CANDIDATE ID, CANDIDATE NAME, NO. OF QUESTIONS; 0 records",
     "NO", "N/A", "None", "—", "PASS"),

    ("Assessments — Qualified sub-tab",
     "Navigate to Qualified Assessments sub-tab",
     "Table loads with qualified records",
     "Loaded; empty state shown; 0 records",
     "NO", "N/A", "None", "—", "PASS"),

    ("Assessments — Disqualified sub-tab",
     "Navigate to Disqualified Assessments sub-tab",
     "Table loads with disqualified records",
     "Loaded; columns: S.NO, ASSESSMENT, INITIATED ON, CANDIDATE ID, CANDIDATE NAME, TEST DURATION, EVALUATED BY, NO. OF QUESTIONS; 0 records",
     "NO", "N/A", "None", "—", "PASS"),

    ("Publish — Job Boards & Career Portal",
     "Open Publish dropdown → confirm 'Job Boards & Career Portal' option exists",
     "Option visible in menu",
     "Option confirmed in Publish dropdown menu",
     "NO", "Menu item visible", "None", "publish_menu.png", "PASS"),

    ("Publish — Social Media",
     "Open Publish → Social Media",
     "Social media publish dialog opens",
     "Dialog 'Publish To' opened with LinkedIn (Configure), Twitter (Configure), Facebook (Share); Cancel + Publish buttons",
     "NO", "Publish To dialog", "None", "publish_social_media.png", "PASS"),

    ("Publish — To Partners",
     "Open Publish → To Partners",
     "Partners list dialog opens",
     "Dialog 'Publish Job to Partners' opened; 2 tabs: All Partners / Published Partners; table with CONTACT NAME, ACCOUNT NAME, EMAIL; 7+ partner contacts listed",
     "NO", "Publish Job to Partners dialog", "None", "publish_to_partners.png", "PASS"),
]

for idx, (area, action, expected, actual, oops, dialog, issues, screenshot, status) in enumerate(rows, 1):
    r = idx + 3
    alt = (idx % 2 == 0)
    ws.row_dimensions[r].height = 60
    data_cell(ws, r, 1, idx, alt, bold=True)
    data_cell(ws, r, 2, area, alt, bold=True)
    data_cell(ws, r, 3, action, alt)
    data_cell(ws, r, 4, expected, alt)
    data_cell(ws, r, 5, actual, alt)
    # OOPS cell
    oc = ws.cell(row=r, column=6, value=oops)
    oc.fill = PatternFill("solid", fgColor=PASS_BG); oc.font = Font(bold=True, color=PASS_FG, size=10)
    oc.alignment = Alignment(horizontal="center", vertical="center"); oc.border = brd()
    data_cell(ws, r, 7, dialog, alt)
    # Issues cell
    ic = ws.cell(row=r, column=8, value=issues)
    if issues != "None":
        ic.fill = PatternFill("solid", fgColor=WARN_BG); ic.font = Font(bold=True, color=WARN_FG, size=10)
    else:
        if alt: ic.fill = PatternFill("solid", fgColor=ALT_BG)
        ic.font = Font(size=10, color="276221")
    ic.alignment = Alignment(vertical="top", wrap_text=True); ic.border = brd()
    data_cell(ws, r, 9, screenshot, alt)
    status_cell(ws, r, 10, status)

# Summary row
r = len(rows) + 4
ws.row_dimensions[r].height = 36
ws.merge_cells(f"A{r}:E{r}")
sc = ws[f"A{r}"]
total = len(rows)
passed = sum(1 for row in rows if row[-1] == "PASS")
info_count = sum(1 for row in rows if row[-1] == "INFO")
warn_count = sum(1 for row in rows if row[-1] == "WARN")
sc.value = f"OVERALL SUMMARY: {total} features tested | {passed} PASS | {info_count} INFO (no candidate) | {warn_count} WARN | 0 FAIL | 0 OOPS errors found across entire Job Posting Module"
sc.font  = Font(bold=True, size=11, color="1F4E79")
sc.fill  = PatternFill("solid", fgColor="DDEBF7")
sc.alignment = Alignment(vertical="center", wrap_text=True)
sc.border = mbrd()

oops_cell = ws.cell(row=r, column=6, value="0 OOPS")
oops_cell.fill = PatternFill("solid", fgColor=PASS_BG); oops_cell.font = Font(bold=True, color=PASS_FG, size=11)
oops_cell.alignment = Alignment(horizontal="center", vertical="center"); oops_cell.border = mbrd()

issues_cell = ws.cell(row=r, column=8, value="3 bugs found (typo, duplicate field, test fields in prod)")
issues_cell.fill = PatternFill("solid", fgColor=WARN_BG); issues_cell.font = Font(bold=True, color=WARN_FG, size=10)
issues_cell.alignment = Alignment(horizontal="center", vertical="center"); issues_cell.border = mbrd()

status_cell(ws, r, 10, "PASS")

# ── Also add a new Bug entry to Bug Report sheet for the typo ──
bug_sheet = wb["Bug Report"] if "Bug Report" in wb.sheetnames else None
if bug_sheet:
    # Find next empty row
    next_row = bug_sheet.max_row + 1
    bug_data = [
        next_row - 3,  # bug number
        "Assessments",
        "Typo: 'Write on Behalf of Candidatee' button label",
        "Button should read 'Write on Behalf of Candidate'",
        "Button label reads 'Candidatee' (extra 'e') on Assessments tab (Oorwin provider)",
        "#/jobs/9554/assessments?type=tagged&tool_type=2",
        "LOW",
        "UI/Cosmetic",
        "assessment_tagged_success.png"
    ]
    for col, val in enumerate(bug_data, 1):
        c = bug_sheet.cell(row=next_row, column=col, value=val)
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = brd()
        if col == 7:  # severity
            c.fill = PatternFill("solid", fgColor="FFF2CC")
            c.font = Font(bold=True, color="9C6500", size=10)

wb.save(r'C:\Oorwin_AI\OorwinLogin.xlsx')
print(f"Updated OorwinLogin.xlsx — added 'Functionality Test' sheet with {len(rows)} test cases")
print(f"Results: {passed} PASS, {info_count} INFO, 0 FAIL, 0 OOPS")
