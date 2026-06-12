import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Oorwin_AI\OorwinLogin.xlsx')

HDR_BG  = "1F3864"; HDR_FG  = "FFFFFF"
PASS_BG = "C6EFCE"; PASS_FG = "276221"
FAIL_BG = "FFE0E0"; FAIL_FG = "9C0006"
WARN_BG = "FFEB9C"; WARN_FG = "9C6500"
ALT_BG  = "F2F2F2"

thin = Side(style="thin", color="BBBBBB")
med  = Side(style="medium", color="888888")
def brd(): return Border(top=thin, bottom=thin, left=thin, right=thin)
def mbrd(): return Border(top=med, bottom=med, left=med, right=med)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=HDR_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = mbrd()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def data_cell(ws, row, col, value, alt=False, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    if alt: c.fill = PatternFill("solid", fgColor=ALT_BG)
    c.font      = Font(bold=bold, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border    = brd()

def status_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    v = (value or "").upper()
    if v == "PASS":   c.fill, c.font = PatternFill("solid", fgColor=PASS_BG), Font(bold=True, color=PASS_FG)
    elif v == "WARN": c.fill, c.font = PatternFill("solid", fgColor=WARN_BG), Font(bold=True, color=WARN_FG)
    elif v == "FAIL": c.fill, c.font = PatternFill("solid", fgColor=FAIL_BG), Font(bold=True, color=FAIL_FG)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = brd()

# ── New sheet: Job Detail Tabs ──────────────────────────────────
if "Job Detail Tabs" in wb.sheetnames:
    del wb["Job Detail Tabs"]

ws = wb.create_sheet("Job Detail Tabs")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

# Banner
ws.merge_cells("A1:J1")
t = ws["A1"]
t.value = "Job Posting Module — Job Detail Tab-by-Tab Test Results  |  Job: OOJ-86071 (ID: 9554)"
t.font  = Font(bold=True, size=13, color=HDR_FG)
t.fill  = PatternFill("solid", fgColor=HDR_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:J2")
sub = ws["A2"]
sub.value = "URL Base: https://oorwinlabs.beanhiredev.com  |  Date: 2026-06-11  |  All tabs tested on Job OOJ-86071"
sub.font  = Font(italic=True, size=10)
sub.fill  = PatternFill("solid", fgColor="D9E1F2")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

headers = ["#", "Tab Name", "URL", "Page Title", "Content / Sections Found",
           "Table Columns Visible", "Empty State Msg", "OOPS?",
           "Console Errors (delta)", "Status"]
widths  = [4,   22,   40,  14,  40,  40,  28,  10,  20,  12]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    hdr_cell(ws, 3, i, h, w)
ws.row_dimensions[3].height = 26

tabs = [
    # (tab, url, title, content, columns, empty_msg, oops, console_delta, status)
    ("1 — Overview",
     "#/jobs/9554/overview",
     "Jobs View",
     "Job header: OOJ-86071 · Active (0 day ago) · Java developer · Dallas Texas, USA · Positions: 1. Stats bar: Pipeline/Submitted/End client/Interviews/Confirmations/Rejected/Onboarded. Full overview layout rendered.",
     "N/A (no table)",
     "N/A",
     "NO", "+0", "PASS"),

    ("2 — Info",
     "#/jobs/9554/info",
     "Jobs View",
     "6 accordion sections expanded: Job Details, Text Field Validation, Pay & Billing Details, Recruitment Team, Recruiter Instructions, Job Description. Attachments and Tags sections also visible.",
     "N/A (accordion view)",
     "N/A",
     "NO", "+3", "PASS"),

    ("3 — Considered Candidates",
     "#/jobs/9554/consideration_details?page=1",
     "Jobs View",
     "Full table rendered with filter_list on each column. 0 records (job has no candidates). Empty state shown.",
     "SUBMISSION ID, CANDIDATE ID, FULL NAME, EMAIL, MOBILE, STATUS, CITY, STATE, COUNTRY, SCORE, FIRST NAME, LAST NAME",
     "No Data\nLooks like you don't have any list to view.",
     "NO", "+0", "PASS"),

    ("4 — Pipeline",
     "#/jobs/9554/pipeline_details/list/-2",
     "Jobs View",
     "Pipeline list view loaded with stage navigation. Table with candidate columns visible. 0 records. 'No Data' empty state displayed correctly.",
     "FULL NAME, YEARS OF EXPERIENCE, JOB TITLE, CITY, STATE, CREATED DATE, STATUS, RATING, OWNER, PIPELINE OWNER",
     "No Data\nLooks like you don't have any list to view.",
     "NO", "+3", "PASS"),

    ("5 — Submissions",
     "#/jobs/9554/submission_details",
     "Jobs View",
     "Submissions table loaded. All column headers visible. 0 records. right_panel_open button visible for side panel. Filter icons on all columns.",
     "SUBMISSION ID, FULL NAME, STATE, CITY, COUNTRY, JOB TITLE, YEARS OF EXPERIENCE, SUBMITTER, AVAILABILITY, STATUS",
     "No records shown (0 submissions)",
     "NO", "+0", "PASS"),

    ("6 — Interviews",
     "#/jobs/9554/interviews",
     "Jobs View",
     "Interviews table loaded with all columns. 0 records. Filter icons on all columns. Table structure intact.",
     "INTERVIEWER, SCHEDULED BY, STATUS, ACCOUNT NAME, TITLE, STAGE, SUBMISSION STAGE, END CLIENT, INTERVIEW RECIPIENT, LOCATION",
     "No records shown (0 interviews)",
     "NO", "+1", "PASS"),

    ("7 — Communications",
     "#/jobs/9554/communications/mail_track",
     "Jobs View",
     "7 communication sub-tabs: Synced Emails (0), Sent Emails (0), Calls (0), SMS (0), Job Shares (0), Follow Up Campaigns (0), Bot Conversation (0), Spark Hire Video Interviews (0). Email connect prompt shown when no email linked. Proceed link to email preferences available.",
     "N/A (sub-tabs layout)",
     "N/A — prompt to connect email account shown",
     "NO", "+2", "PASS"),

    ("8 — Recruiter Actions",
     "#/jobs/9554/settings",
     "Jobs View",
     "Recruiter Actions / Settings table loaded. Job board columns: RECRUITERS, TWITTER, LINKEDIN, DICE, MONSTER, CAREER BUILDER, TECH FETCH, SNAPRECRUIT, EZJOBS, NEXXT, INDEED, POST JOB FREE, ZIPRECRUITER, JOBRAPIDO, GLASSDOOR, JOOBLE, CAREERJET, JOBOMAS, SHEJOBS, MONSTERFREE, JORA, RECRUIT, TALENT, DR. JOBS PRO, PROMOTION MAILS, SUBMISSIONS. Track Status, Clear, Save buttons present.",
     "RECRUITERS, TWITTER, LINKEDIN, DICE, MONSTER, CAREER BUILDER, TECH FETCH, SNAPRECRUIT, EZJOBS, NEXXT, INDEED + more",
     "N/A",
     "NO", "+7 (⚠ highest delta)", "WARN"),

    ("9 — Assessments",
     "#/jobs/9554/assessments?type=tagged",
     "Jobs View",
     "Assessment sub-tabs: Tagged Assessments, Initiated Assessments (0), Pending (0), Qualified (0), Disqualified (0). Provider tabs: Adaface, Oorwin, Testlify, Codility. Tag Assessments and Initiate Assessment buttons present. Table loaded.",
     "S.NO, ASSESSMENT, TAGGED ON, TEST DURATION, CUT-OFF SCORE, TAGGED BY",
     "No Data",
     "NO", "+2", "PASS"),
]

for idx, (tab, url, title, content, cols, empty, oops, delta, status) in enumerate(tabs, 1):
    r = idx + 3
    alt = (idx % 2 == 0)
    ws.row_dimensions[r].height = 70
    data_cell(ws, r, 1, idx, alt, bold=True)
    data_cell(ws, r, 2, tab, alt, bold=True)
    data_cell(ws, r, 3, url, alt)
    data_cell(ws, r, 4, title, alt)
    data_cell(ws, r, 5, content, alt)
    data_cell(ws, r, 6, cols, alt)
    data_cell(ws, r, 7, empty, alt)
    # OOPS cell
    oc = ws.cell(row=r, column=8, value=oops)
    oc.fill = PatternFill("solid", fgColor=PASS_BG); oc.font = Font(bold=True, color=PASS_FG)
    oc.alignment = Alignment(horizontal="center", vertical="center"); oc.border = brd()
    # Console delta
    cc = ws.cell(row=r, column=9, value=delta)
    if "⚠" in delta or int(delta.replace("+","").replace("⚠ highest delta","").strip() or "0") >= 5:
        cc.fill = PatternFill("solid", fgColor=WARN_BG); cc.font = Font(bold=True, color=WARN_FG)
    else:
        cc.fill = PatternFill("solid", fgColor="F2F2F2")
    cc.alignment = Alignment(horizontal="center", vertical="center"); cc.border = brd()
    status_cell(ws, r, 10, status)

# Summary row
r = len(tabs) + 4
ws.row_dimensions[r].height = 30
ws.merge_cells(f"A{r}:G{r}")
sc = ws[f"A{r}"]
sc.value = "OVERALL: All 9 tabs loaded successfully. No OOPS errors. No blank/broken pages. Console errors accumulated to 46 total (+9 new across last 3 tabs — Recruiter Actions contributed +7)."
sc.font  = Font(bold=True, size=10, color="1F4E79")
sc.fill  = PatternFill("solid", fgColor="DDEBF7")
sc.alignment = Alignment(vertical="center", wrap_text=True)
sc.border = mbrd()

sc2 = ws.cell(row=r, column=8, value="NO")
sc2.fill = PatternFill("solid", fgColor=PASS_BG); sc2.font = Font(bold=True, color=PASS_FG)
sc2.alignment = Alignment(horizontal="center", vertical="center"); sc2.border = mbrd()

sc3 = ws.cell(row=r, column=9, value="+46 total")
sc3.fill = PatternFill("solid", fgColor=WARN_BG); sc3.font = Font(bold=True, color=WARN_FG)
sc3.alignment = Alignment(horizontal="center", vertical="center"); sc3.border = mbrd()

sc4 = ws.cell(row=r, column=10, value="8 PASS / 1 WARN")
sc4.fill = PatternFill("solid", fgColor=PASS_BG); sc4.font = Font(bold=True, color=PASS_FG)
sc4.alignment = Alignment(horizontal="center", vertical="center"); sc4.border = mbrd()

wb.save(r'C:\Oorwin_AI\OorwinLogin.xlsx')
print("Updated: OorwinLogin.xlsx — added 'Job Detail Tabs' sheet")
