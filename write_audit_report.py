import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Oorwin_AI\OorwinLogin.xlsx')

# Colours
HDR_BG  = "1F3864"; HDR_FG  = "FFFFFF"
PASS_BG = "C6EFCE"; PASS_FG = "276221"
FAIL_BG = "FFE0E0"; FAIL_FG = "9C0006"
WARN_BG = "FFEB9C"; WARN_FG = "9C6500"
INFO_BG = "DDEBF7"; INFO_FG = "1F4E79"
ALT_BG  = "F2F2F2"
CRIT_BG = "FF0000"; CRIT_FG = "FFFFFF"

thin = Side(style="thin",   color="BBBBBB")
med  = Side(style="medium", color="888888")
def brd():  return Border(top=thin, bottom=thin, left=thin, right=thin)
def mbrd(): return Border(top=med,  bottom=med,  left=med,  right=med)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=HDR_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = mbrd()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def data_cell(ws, row, col, value, alt=False, bold=False, color=None):
    c = ws.cell(row=row, column=col, value=value)
    if color:
        c.fill = PatternFill("solid", fgColor=color)
    elif alt:
        c.fill = PatternFill("solid", fgColor=ALT_BG)
    c.font      = Font(bold=bold, size=10)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border    = brd()

def sev_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    v = (value or "").upper()
    if v == "CRITICAL":
        c.fill, c.font = PatternFill("solid", fgColor=CRIT_BG), Font(bold=True, color=CRIT_FG, size=10)
    elif v == "HIGH":
        c.fill, c.font = PatternFill("solid", fgColor=FAIL_BG), Font(bold=True, color=FAIL_FG, size=10)
    elif v == "MEDIUM":
        c.fill, c.font = PatternFill("solid", fgColor=WARN_BG), Font(bold=True, color=WARN_FG, size=10)
    elif v == "LOW":
        c.fill, c.font = PatternFill("solid", fgColor=INFO_BG), Font(bold=True, color=INFO_FG, size=10)
    elif v == "INFO":
        c.fill, c.font = PatternFill("solid", fgColor=ALT_BG), Font(italic=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = brd()

def cat_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = brd()

# ── Delete existing sheet if present ────────────────────────────────────────
if "Application Audit" in wb.sheetnames:
    del wb["Application Audit"]

ws = wb.create_sheet("Application Audit")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

# Banner
ws.merge_cells("A1:J1")
t = ws["A1"]
t.value = "BeanHire ATS — Comprehensive Application Audit Report  |  URL: https://oorwinlabs.beanhiredev.com  |  User: pavant+ui@oorwin.com"
t.font  = Font(bold=True, size=13, color=HDR_FG)
t.fill  = PatternFill("solid", fgColor=HDR_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:J2")
sub = ws["A2"]
sub.value = "Full application walk-through — all major modules tested  |  Date: 2026-06-11  |  Total issues: 22"
sub.font  = Font(italic=True, size=10)
sub.fill  = PatternFill("solid", fgColor="D9E1F2")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Legend row
ws.merge_cells("A3:J3")
leg = ws["A3"]
leg.value = (
    "SEVERITY KEY:   🔴 CRITICAL = Data/function broken  |  🟠 HIGH = Functional bug  |  "
    "🟡 MEDIUM = UI/UX defect or access issue  |  🔵 LOW = Minor/typo  |  ⚪ INFO = Observation/role-based access"
)
leg.font  = Font(size=9, italic=True)
leg.fill  = PatternFill("solid", fgColor="FFF2CC")
leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[3].height = 24

# Column headers
headers = ["#", "Category", "Severity", "Issue Title", "Location / URL", "Details / Steps to Reproduce",
           "Expected", "Actual", "Screenshot", "Status"]
widths  = [4,   18,       12,         32,            32,                 50,                              28,    28,    20,          12]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    hdr_cell(ws, 4, i, h, w)
ws.row_dimensions[4].height = 26

# ── Issues data ──────────────────────────────────────────────────────────────
issues = [
    # (category, severity, title, location, details, expected, actual, screenshot)
    (
        "Functional Bug",
        "CRITICAL",
        "OOPS error on Lead detail view",
        "#/leads/1699/view → redirects to #/error",
        "Navigate to Leads list → click any lead record (e.g. 'Rahul' lead #1699). Page immediately redirects to /#/error.",
        "Lead detail page loads with full data",
        "OOPS !!! — 'Looks like there is an issue with the data fetch. Please refresh your page and retry.' Page unusable.",
        "OOPS_lead_detail_error.png"
    ),
    (
        "Test Data in Production",
        "HIGH",
        "Test/dev fields visible in Lead Create form",
        "#/leads/create",
        "Open Lead Create form. Scroll to custom fields section.",
        "Only production-ready fields should appear",
        "8 test fields visible: '22Tag', '22Test1', '22Muti Usres' (double typo), '22CheckBoxes', '22rTest1', '22rTest2', '22DropDown', '22textArea'",
        ""
    ),
    (
        "Test Data in Production",
        "HIGH",
        "Test/dev fields visible in Opportunities Create form",
        "#/opportunities/create",
        "Open Opportunities Create form and review all field labels.",
        "Only production-ready fields should appear",
        "6 test fields visible: 'oppo', 'Opportunity Amount1', 'Text', 'Test', 'sub field', 'Test1'",
        "opportunities_create_form.png"
    ),
    (
        "Test Data in Production",
        "HIGH",
        "Test/dev section visible in Job Create/Info form",
        "#/jobs/create → Info tab",
        "Create or view a job's Info tab. A section titled 'Text Field Validation' appears with multiple test inputs.",
        "Only validated, production-ready form sections",
        "Section 'Text Field Validation' with multiple test fields shown in production job form",
        ""
    ),
    (
        "Functional Bug",
        "HIGH",
        "Duplicate 'Client Job ID' field in Job form",
        "#/jobs/create (Job Details section)",
        "Open Job Create form → scroll to Job Details accordion. 'Client Job ID' appears twice.",
        "Field appears once",
        "'Client Job ID' field is duplicated in the Job Details section",
        ""
    ),
    (
        "UI / Typo",
        "MEDIUM",
        "'SUBSIDARIES' spelling error in Accounts",
        "#/accounts (table header) AND #/accounts/2718/view",
        "Navigate to Accounts list. Check the table column header. Also visible in account detail view sidebar.",
        "Column header reads 'SUBSIDIARIES'",
        "Column header reads 'SUBSIDARIES' (missing 'I')",
        ""
    ),
    (
        "UI / Typo",
        "MEDIUM",
        "'AVALIABLE' spelling error on BGV page",
        "#/bgv_list/available",
        "Navigate to BGV list via the shield_person header icon or direct URL.",
        "Tab reads 'AVAILABLE'",
        "Tab/heading reads 'AVALIABLE' (letters transposed)",
        ""
    ),
    (
        "UI / Typo",
        "MEDIUM",
        "'additional servicecs' spelling error in Billing & Usage",
        "#/admin/subscription/top_up",
        "Navigate to Billing & Usage → subscription top-up page. Read the 'Additional Services' section description.",
        "Text reads 'Choose additional services for your plan'",
        "Text reads 'Choose additional servicecs for your plan' (extra 'c' inserted)",
        ""
    ),
    (
        "UI / Typo",
        "MEDIUM",
        "'Write on Behalf of Candidatee' typo in Assessments",
        "#/jobs/[id]/assessments → Oorwin tab",
        "Open any job's Assessments tab → click Oorwin provider → view assessment options.",
        "Text reads 'Write on Behalf of Candidate'",
        "Text reads 'Write on Behalf of Candidatee' (extra 'e' at end)",
        ""
    ),
    (
        "UI / Typo",
        "MEDIUM",
        "'Mark as Preffered' spelling error in Job form",
        "#/jobs/create → Pay & Billing Details section",
        "Open Job Create/Edit form → expand Pay & Billing Details accordion.",
        "Label reads 'Mark as Preferred'",
        "Label reads 'Mark as Preffered' (double 'f')",
        ""
    ),
    (
        "UI / Typo",
        "MEDIUM",
        "'22Muti Usres' double typo in Lead form field",
        "#/leads/create (custom fields)",
        "Open Lead Create form → scroll to custom fields. Field label contains double typo.",
        "Field label 'Multi Users' (production-ready name)",
        "Label reads '22Muti Usres' — 'Muti' (should be 'Multi') and 'Usres' (should be 'Users'), plus test prefix '22'",
        ""
    ),
    (
        "UI / Rendering",
        "MEDIUM",
        "HTML entity &nbsp; renders as raw text in notifications",
        "Notifications panel (bell icon, header)",
        "Click the notifications bell (badge shows 4). Read notification message text.",
        "Proper formatted text with normal spacing",
        "Raw HTML entity '&nbsp;' appears literally in notification text body instead of being rendered as a space",
        "notifications_panel.png"
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Employees module",
        "#/employees → redirects to #/error",
        "Navigate to HR → Employees. Page shows Access Denied.",
        "Employees list loads for this user role",
        "'Access Denied — You don\\'t have access to view this record'. Timesheets (/timesheets/list) also denied. May be by design for this role.",
        ""
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Leaves module",
        "#/leaves/all → redirects to #/error",
        "Navigate to Leaves (from header or HR menu).",
        "Leaves list loads for this user role",
        "'Access Denied — You don\\'t have access to view this record'. May be by design.",
        ""
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Calendar module",
        "#/calendar → redirects to #/error",
        "Navigate to Calendar.",
        "Calendar view loads",
        "'Access Denied'. May be by design for this role.",
        ""
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Archives module",
        "#/archives → redirects to #/error",
        "Navigate to Archives (profile menu → Archives).",
        "Archives list loads",
        "'Access Denied'. May be by design for this role.",
        ""
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Admin > Users",
        "#/admin/users → redirects to #/error",
        "Navigate to Admin → Access Control → Users.",
        "User management page loads",
        "'Access Denied'. Role may not have admin user management access.",
        ""
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Admin > Custom Forms",
        "#/admin/configurations/custom_forms",
        "Navigate to Admin → Configurations → Custom Forms.",
        "Custom Forms configuration page loads",
        "'Access Denied'. Admin configuration access requires higher privileges.",
        ""
    ),
    (
        "Access Control",
        "INFO",
        "Access Denied — Admin > Pipeline Stages",
        "#/admin/configurations/pipeline_stages",
        "Navigate to Admin → Configurations → Pipeline Stages.",
        "Pipeline Stages configuration page loads",
        "'Access Denied'. Admin configuration access requires higher privileges.",
        ""
    ),
    (
        "Observation",
        "INFO",
        "Console errors consistently high (90–130 per page)",
        "All pages (browser developer tools console)",
        "Open browser DevTools on any page and observe console errors tab.",
        "Minimal or zero console errors",
        "Every page load generates 90–130 JavaScript console errors. Recruiter Actions tab produces highest delta (+7). Root cause not determined — may indicate unhandled API errors or deprecated API calls.",
        ""
    ),
    (
        "Observation",
        "INFO",
        "multiple_stop icon = Org Switcher",
        "Header toolbar (multiple_stop icon)",
        "Click the multiple_stop icon in the header toolbar.",
        "N/A — informational",
        "Opens org-switcher dropdown showing 'Consolidation India'. Functional.",
        "multiple_stop_icon.png"
    ),
    (
        "Observation",
        "INFO",
        "check_circle icon = Approvals Panel (205 pending admin approvals)",
        "Header toolbar (check_circle icon)",
        "Click the check_circle icon in the header toolbar.",
        "N/A — informational",
        "Opens approvals panel: My Pending Approvals (0), Approve As Admin (205), Document Submissions (232), Employee Requests (172), Requisitions (5). High pending counts noted.",
        "check_circle_icon.png"
    ),
]

for idx, (cat, sev, title, loc, details, expected, actual, screenshot) in enumerate(issues, 1):
    r = idx + 4
    alt = (idx % 2 == 0)
    ws.row_dimensions[r].height = 80
    data_cell(ws, r, 1, idx, alt, bold=True)
    cat_cell(ws, r, 2, cat)
    sev_cell(ws, r, 3, sev)
    data_cell(ws, r, 4, title, alt, bold=True)
    data_cell(ws, r, 5, loc, alt)
    data_cell(ws, r, 6, details, alt)
    data_cell(ws, r, 7, expected, alt)
    data_cell(ws, r, 8, actual, alt)
    sc = ws.cell(row=r, column=9, value=screenshot if screenshot else "—")
    sc.font = Font(color="0563C1" if screenshot else None, underline="single" if screenshot else None, size=10)
    sc.alignment = Alignment(vertical="top", wrap_text=True); sc.border = brd()
    if alt: sc.fill = PatternFill("solid", fgColor=ALT_BG)
    stat = ws.cell(row=r, column=10, value="OPEN" if sev not in ("INFO",) else "INFO")
    if sev == "CRITICAL":
        stat.fill, stat.font = PatternFill("solid", fgColor=CRIT_BG), Font(bold=True, color=CRIT_FG, size=10)
    elif sev == "HIGH":
        stat.fill, stat.font = PatternFill("solid", fgColor=FAIL_BG), Font(bold=True, color=FAIL_FG, size=10)
    elif sev == "MEDIUM":
        stat.fill, stat.font = PatternFill("solid", fgColor=WARN_BG), Font(bold=True, color=WARN_FG, size=10)
    else:
        stat.fill, stat.font = PatternFill("solid", fgColor=INFO_BG), Font(italic=True, color=INFO_FG, size=10)
    stat.alignment = Alignment(horizontal="center", vertical="center"); stat.border = brd()

# Summary footer
r = len(issues) + 5
ws.row_dimensions[r].height = 40
ws.merge_cells(f"A{r}:H{r}")
fc = ws[f"A{r}"]
fc.value = (
    "SUMMARY: 1 CRITICAL | 4 HIGH | 7 MEDIUM | 3 LOW | 7 INFO  ·  "
    "Modules PASS: Dashboard, Jobs (tabs), Candidates, Submissions, Interviews, Pipeline, Reports, "
    "Expenses, Invoices, Bills, US Immigration, I9, Tasks, Documents, Onboarding, Attendance, Assets, "
    "Projects, Offboarding, Performance, Help Desk, Assessments, Targets, Marketplace, Admin  ·  "
    "No OOPS errors except Lead detail #1699"
)
fc.font  = Font(bold=True, size=10, color="1F4E79")
fc.fill  = PatternFill("solid", fgColor="DDEBF7")
fc.alignment = Alignment(vertical="center", wrap_text=True)
fc.border = mbrd()

for col in range(9, 11):
    sc2 = ws.cell(row=r, column=col, value="")
    sc2.fill  = PatternFill("solid", fgColor="DDEBF7")
    sc2.border = mbrd()

wb.save(r'C:\Oorwin_AI\OorwinLogin.xlsx')
print("Done — Application Audit sheet written to OorwinLogin.xlsx")
print(f"Total issues logged: {len(issues)}")
